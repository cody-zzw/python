# -*- coding: UTF-8 -*-
#!/usr/bin/env python3
import sys
import os
import time
import configparser
from tkinter import *
from tkinter import ttk
import openpyxl
from openpyxl.styles import Font, colors, Alignment
from decimal import Decimal   #浮点处理

class Excel():
	def __init__(self,my_database):
		self.my_database = my_database
		
	def create_month_income_excel(self,year,month):
		wookbook = openpyxl.Workbook()
		sheet1 = wookbook.create_sheet(title="sheet1",index=0)
		path_month = os.path.join(self.my_database.path_income,year,month)
		excel_name = year+"年"+month+"月"+"计件表"
		excel_path = os.path.join(self.my_database.path_excel,excel_name+".xlsx")
		if os.path.exists(excel_path):
			os.remove(excel_path)
		sheet1.column_dimensions['B'].width = 25
		sheet1.column_dimensions['C'].width = 25
		sheet1.column_dimensions['I'].width = 20
		sheet1['A1'].value = excel_name
		sheet1['A1'].alignment = Alignment(horizontal='center', vertical='center')
		sheet1.merge_cells(range_string='A1:I1')
		sheet1.append(["日期","料号","型号","数量","折盒单价","折盒金额","单价","金额","送货单号"])
		day_list = os.listdir(path_month)
		for day in day_list:
			path_day = os.path.join(path_month,day)
			data = self.my_database.read_incone_data(path_day)
			for data_child in data:
				data_child.insert(0,day)
				data_child[3] = Decimal(data_child[3])
				data_child[4] = Decimal(data_child[4])
				data_child[5] = Decimal(data_child[5])
				data_child[6] = Decimal(data_child[6])
				data_child[7] = Decimal(data_child[7])
				sheet1.append(data_child)
		
		wookbook.save(excel_path)
		
	def create_month_expenditure_excel(self,year,month):
		wookbook = openpyxl.Workbook()
		sheet1 = wookbook.create_sheet(title="sheet1",index=0)
		path_month = os.path.join(self.my_database.path_expenditure,year,month)
		excel_name = year+"年"+month+"月"+"支出表"
		excel_path = os.path.join(self.my_database.path_excel,excel_name+".xlsx")
		if os.path.exists(excel_path):
			os.remove(excel_path)
		sheet1.column_dimensions['B'].width = 15
		sheet1.column_dimensions['G'].width = 20
		sheet1['A1'].value = excel_name
		sheet1['A1'].alignment = Alignment(horizontal='center', vertical='center')
		sheet1.merge_cells(range_string='A1:G1')
		sheet1.append(["日期","姓名","工价","工时","金额","补贴","联系方式"])
		day_list = os.listdir(path_month)
		for day in day_list:
			path_day = os.path.join(path_month,day)
			data = self.my_database.read_expenditure_data(path_day)
			for data_child in data:
				data_child.insert(0,day)
				data_child[2] = Decimal(data_child[2])
				data_child[3] = Decimal(data_child[3])
				data_child[4] = Decimal(data_child[4])
				data_child[5] = Decimal(data_child[5])
				sheet1.append(data_child)
		
		wookbook.save(excel_path)



class Database():
	def __init__(self):
		"""初始化数据库"""
		print(1)
		self.tm_year = time.localtime(time.time()).tm_year  #获取当前年份
		self.tm_mon = time.localtime(time.time()).tm_mon  #获取当前月
		self.tm_mday = time.localtime(time.time()).tm_mday  #获取当前日
		self.pwd = os.getcwd()  #当前工作目录
		self.path_database = os.path.join(self.pwd,"database")
		self.path_income = os.path.join(self.path_database,"income")
		self.path_expenditure = os.path.join(self.path_database,"expenditure")
		self.path_part_No = os.path.join(self.path_database,"Part_No.cfg")
		self.path_name = os.path.join(self.path_database,"name.cfg")
		self.path_excel = os.path.join(self.pwd,"excel")
		if not os.path.exists(self.path_database):
			os.mkdir(self.path_database)
		if not os.path.exists(self.path_income):
			os.mkdir(self.path_income)
		if not os.path.exists(self.path_expenditure):
			os.mkdir(self.path_expenditure)
		if not os.path.exists(self.path_excel):
			os.mkdir(self.path_excel)
		
		#创建今年所有月份目录
		path_income_year = os.path.join(self.path_income,str(self.tm_year))
		if not os.path.exists(path_income_year):
			os.mkdir(path_income_year)
			for i in range(1,13):
				path_income_mon = os.path.join(path_income_year,str(i))
				os.mkdir(path_income_mon)
		path_expenditure_year = os.path.join(self.path_expenditure,str(self.tm_year))
		if not os.path.exists(path_expenditure_year):
			os.mkdir(path_expenditure_year)
			for i in range(1,13):
				path_expenditure_mon = os.path.join(path_expenditure_year,str(i))
				os.mkdir(path_expenditure_mon)
				
	def create_part_No_file(self):
		raw_file_path = os.path.join(self.path_database,"raw.txt")
		with open(raw_file_path,"r",encoding="utf-8") as f:
			line_list = f.readlines()
		node = line_list.index("####\n")
		part_No_list = line_list[:node]
		item_type_list = line_list[node+1:]
		part_No_list_new = []
		for x in part_No_list:
			if x not in part_No_list_new:
				part_No_list_new.append(x)
		
		conf = configparser.RawConfigParser()
		i = 1
		for part_No in part_No_list_new:
			node = part_No_list.index(part_No)
			conf.add_section(str(i))
			conf.set(str(i),"part_No",part_No[:len(part_No)-1])
			conf.set(str(i),"type",item_type_list[node][:len(item_type_list[node])-1])
			i += 1
		with open(self.path_part_No,'w',encoding='utf-8') as f:
			conf.write(f)
			
	def get_year_list(self):
		"""获取数据库所有年份"""
		dirs = os.listdir(self.path_income)
		for i in range(0,len(dirs)):
			dirs[i] = int(dirs[i])
		return dirs
		
	def read_incone_data(self,file_path):
		if not os.path.exists(file_path):
			return None
		conf = configparser.RawConfigParser()
		conf.read(file_path,encoding='utf-8')
		data = []
		all_sections = conf.sections()
		if len(all_sections) == 0:
			return None
		for section in all_sections:
			data_child = []
			data_child.append(conf.get(section,"part_No"))
			data_child.append(conf.get(section,"type"))
			data_child.append(conf.get(section,"number"))
			data_child.append(conf.get(section,"folding_box_unit_price"))
			data_child.append(conf.get(section,"folding_box_total_price"))
			data_child.append(conf.get(section,"unit_price"))
			data_child.append(conf.get(section,"total_price"))
			data_child.append(conf.get(section,"delivery_note_number"))
			data.append(data_child)
		return data
		
	def read_expenditure_data(self,file_path):
		if not os.path.exists(file_path):
			return None
		conf = configparser.RawConfigParser()
		conf.read(file_path,encoding='utf-8')
		data = []
		all_sections = conf.sections()
		if len(all_sections) == 0:
			return None
		for section in all_sections:
			data_child = []
			data_child.append(conf.get(section,"name"))
			data_child.append(conf.get(section,"wages"))
			data_child.append(conf.get(section,"time"))
			data_child.append(conf.get(section,"amount"))
			data_child.append(conf.get(section,"subsidy"))
			data_child.append(conf.get(section,"information"))
			data.append(data_child)
		return data
		
	def write_expenditure_data(self,file_path,data):
		conf = configparser.RawConfigParser()
		for i in range(1,len(data)+1):
			conf.add_section(str(i))
			conf.set(str(i),"name",data[i-1][0])
			conf.set(str(i),"wages",data[i-1][1])
			conf.set(str(i),"time",data[i-1][2])
			conf.set(str(i),"amount",data[i-1][3])
			conf.set(str(i),"subsidy",data[i-1][4])
			conf.set(str(i),"information",data[i-1][5])
		with open(file_path,"w",encoding="utf-8") as f:
			conf.write(f)
	
	def write_income_data(self,file_path,data):
		conf = configparser.RawConfigParser()
		for i in range(1,len(data)+1):
			conf.add_section(str(i))
			conf.set(str(i),"part_No",data[i-1][0])
			conf.set(str(i),"type",data[i-1][1])
			conf.set(str(i),"number",data[i-1][2])
			conf.set(str(i),"folding_box_unit_price",data[i-1][3])
			conf.set(str(i),"folding_box_total_price",data[i-1][4])
			conf.set(str(i),"unit_price",data[i-1][5])
			conf.set(str(i),"total_price",data[i-1][6])
			conf.set(str(i),"delivery_note_number",data[i-1][7])
		with open(file_path,"w",encoding="utf-8") as f:
			conf.write(f)
		
	def get_part_No_list(self):
		conf = configparser.RawConfigParser()
		conf.read(self.path_part_No,encoding="utf-8")
		part_No_list = []
		all_sections = conf.sections()
		for section in all_sections:
			part_No_list.append(conf.get(section,"part_No"))
		return part_No_list
		
	def get_name_list(self):
		conf = configparser.RawConfigParser()
		conf.read(self.path_name,encoding="utf-8")
		name_list = []
		all_sections = conf.sections()
		for section in all_sections:
			name_list.append(conf.get(section,"name"))
		return name_list
		
	def check_part_No(self,part_No):
		conf = configparser.RawConfigParser()
		conf.read(self.path_part_No,encoding="utf-8")
		all_sections = conf.sections()
		for section in all_sections:
			if part_No == conf.get(section,"part_No"):
				item_type = conf.get(section,"type")
				return item_type
		return None
	
	def check_name(self,name):
		conf = configparser.RawConfigParser()
		conf.read(self.path_name,encoding="utf-8")
		all_sections = conf.sections()
		for section in all_sections:
			if name == conf.get(section,"name"):
				wages = conf.get(section,"wages")
				return wages
		return None
	
	def add_name(self,name,wages):
		conf = configparser.RawConfigParser()
		conf.read(self.path_name,encoding="utf-8")
		all_sections = conf.sections()
		new_section = str(len(all_sections)+1)
		conf.add_section(new_section)
		conf.set(new_section,"name",name)
		conf.set(new_section,"wages",wages)
		with open(self.path_name,'w',encoding='utf-8') as f:
			conf.write(f)
		
class GUI():
	def __init__(self,my_database,my_excel):
		"""初始化窗口"""
		self.my_database = my_database
		self.my_excel = my_excel
		self.database_year_list = my_database.get_year_list()
		self.part_No_list = self.my_database.get_part_No_list()
		self.name_list = self.my_database.get_name_list()
		
		self.mon_list = []
		for i in range(1,13):
			self.mon_list.append(i)
		self.mday_list = []
		for j in range(1,self.get_day_num(self.my_database.tm_year,self.my_database.tm_mon)+1):
			self.mday_list.append(j)
			
		self.top = Tk()
		self.top.title("MyBook")
		self.top.geometry("1400x568+0+50")
	
	def get_day_num(self,year,mon):
		"""计算某年某月有多少天"""
		if mon in [1,2,5,7,8,10,12]:
			return 31
		elif mon in [4,6,9,11]:
			return 30
		elif mon == 2 and ((year%4 == 0 and year%100 != 0) or (year%400 == 0)):
			return 29
		else:
			return 28
	
	def top_destory(self):
		"""清除窗口所有控件"""
		for widget in self.top.winfo_children():
			widget.destroy()
		time.sleep(0.1)
	
	def create_drop_down_list(self,select_list):
		"""创建下拉列表"""
		comvalue = StringVar() #窗体自带的文本，新建一个值
		comboxlist = ttk.Combobox(self.top,textvariable=comvalue) #初始化
		comboxlist["values"] = select_list
		return comboxlist
		
	def main_win(self):
		"""主界面"""
		self.top_destory()
		Label(self.top,text = "今日日期："+str(self.my_database.tm_year)+"年"+str(self.my_database.tm_mon)+"月"+str(self.my_database.tm_mday)+"日",width = 30,height = 4).pack()
		Label(self.top,text = "请选择以下项目").pack()
		Button(self.top,text = "收入",width = 30,height = 4,command = self.income_win).pack()
		Button(self.top,text = "支出",width = 30,height = 4,command = self.expenditure_win).pack()
		Button(self.top,text = "年度报表",width = 30,height = 4,command = self.year_report).pack()
		Button(self.top,text = "退出",width = 30,height = 4,command = self.top.quit).pack()
		
	def income_win(self):
		self.state = "收入"
		self.income_and_expenditure_select_date(self.state)
		
	def expenditure_win(self):
		self.state = "支出"
		self.income_and_expenditure_select_date(self.state)
		
	def year_report(self):
		pass
		
	def income_and_expenditure_select_date(self,info):
		self.top_destory()
		Label(self.top,text = "日期选择",width = 30,height = 2).pack(pady=10)
		self.comboxlist_year = self.create_drop_down_list(self.database_year_list)
		self.comboxlist_year.bind("<<ComboboxSelected>>",None)  #绑定事件,下拉列表框被选中时，绑定函数
		self.comboxlist_year.current(self.database_year_list.index(self.my_database.tm_year))  
		self.comboxlist_year.pack(pady=10)
		
		self.comboxlist_mon = self.create_drop_down_list(self.mon_list)
		self.comboxlist_mon.bind("<<ComboboxSelected>>",None)  #绑定事件,下拉列表框被选中时，绑定函数
		self.comboxlist_mon.current(self.mon_list.index(self.my_database.tm_mon))  
		self.comboxlist_mon.pack(pady=10)
		
		self.comboxlist_mday = self.create_drop_down_list(self.mday_list)
		self.comboxlist_mday.bind("<<ComboboxSelected>>",None)  #绑定事件,下拉列表框被选中时，绑定函数
		self.comboxlist_mday.current(self.mday_list.index(self.my_database.tm_mday))  
		self.comboxlist_mday.pack(pady=10)
		
		Button(self.top,text = "生成"+info+"Excel",width = 30,height = 2,command = self.generate_excel).pack(pady=10)
		Button(self.top,text = "确定",width = 30,height = 2,command = self.day_operate).pack(pady=10)
		Button(self.top,text = "返回",width = 30,height = 2,command = self.main_win).pack(pady=10)
		
	def get_select_date(self):
		self.select_year = self.comboxlist_year.get()
		self.select_month = self.comboxlist_mon.get()
		self.select_mday = self.comboxlist_mday.get()
		
	def generate_excel(self):
		self.get_select_date()
		if self.state == "收入":
			self.my_excel.create_month_income_excel(self.select_year,self.select_month)
		else:
			self.my_excel.create_month_expenditure_excel(self.select_year,self.select_month)
	
	def day_operate(self):
		self.get_select_date()
		self.top_destory()
		self.show_content_title()
		
		if self.state == "收入":
			self.show_content_income_item()
			file_path = os.path.join(self.my_database.path_income,str(self.select_year),str(self.select_month),str(self.select_mday))
			self.data = self.my_database.read_incone_data(file_path)
			y = self.show_content_income_detail()
			self.show_content_commit(y)
		else:
			self.show_content_expenditure_item()
			file_path = os.path.join(self.my_database.path_expenditure,str(self.select_year),str(self.select_month),str(self.select_mday))
			self.data = self.my_database.read_expenditure_data(file_path)
			y = self.show_content_expenditure_detail()
			self.show_content_commit(y)
		
	def show_content_title(self):
		lable = Label(self.top,text = str(self.select_year)+"年"+str(self.select_month)+"月"+str(self.select_mday)+"日"+self.state+"明细",width = 20,height = 1)
		if self.state == "收入":
			lable.grid(row=0,column=4)
		else:
			lable.grid(row=0,column=3)
	
	def show_content_expenditure_item(self):
		Label(self.top,text = "序号",width = 16).grid(row=1,column=0)
		Label(self.top,text = "姓名",width = 16).grid(row=1,column=1)
		Label(self.top,text = "工价",width = 16).grid(row=1,column=2)
		Label(self.top,text = "工时",width = 16).grid(row=1,column=3)
		Label(self.top,text = "金额",width = 16).grid(row=1,column=4)
		Label(self.top,text = "补贴",width = 16).grid(row=1,column=5)
		Label(self.top,text = "联系方式",width = 16).grid(row=1,column=6)
		
	
	def show_content_income_item(self):
		Label(self.top,text = "序号",width = 16).grid(row=1,column=0)
		Label(self.top,text = "料号",width = 16).grid(row=1,column=1)
		Label(self.top,text = "型号",width = 16).grid(row=1,column=2)
		Label(self.top,text = "数量",width = 16).grid(row=1,column=3)
		Label(self.top,text = "折盒单价",width = 16).grid(row=1,column=4)
		Label(self.top,text = "折盒金额",width = 16).grid(row=1,column=5)
		Label(self.top,text = "单价",width = 16).grid(row=1,column=6)
		Label(self.top,text = "金额",width = 16).grid(row=1,column=7)
		Label(self.top,text = "送货单号",width = 16).grid(row=1,column=8)
		
	def show_content_expenditure_detail(self):
		if self.data == None:
			self.data = []
			return 2
		self.comboxlist_name_list = []
		self.entry_list = []
		i = 0
		for i in range(1,len(self.data)+1):
			entry_child_list = []
			Label(self.top,text = str(i),width = 16).grid(row=1+i,column=0)
			name_list_tmp = self.name_list[:]
			if self.data[i-1][0] not in name_list_tmp:
				name_list_tmp.append(self.data[i-1][0])
			comboxlist_name = self.create_drop_down_list(name_list_tmp)
			comboxlist_name.bind("<<ComboboxSelected>>",self.refresh)  #绑定事件,下拉列表框被选中时，绑定函数
			comboxlist_name.current(name_list_tmp.index(self.data[i-1][0]))
			comboxlist_name.grid(row=1+i,column=1)
			self.comboxlist_name_list.append(comboxlist_name)
			e = Entry(self.top)
			e.insert(0,str(self.data[i-1][1]))
			e.grid(row=1+i,column=2)
			entry_child_list.append(e)
			e = Entry(self.top)
			e.insert(0,str(self.data[i-1][2]))
			e.grid(row=1+i,column=3)
			entry_child_list.append(e)
			e = Entry(self.top)
			e.insert(0,str(self.data[i-1][3]))
			e.grid(row=1+i,column=4)
			entry_child_list.append(e)
			e = Entry(self.top)
			e.insert(0,str(self.data[i-1][4]))
			e.grid(row=1+i,column=5)
			entry_child_list.append(e)
			e = Entry(self.top)
			e.insert(0,str(self.data[i-1][5]))
			e.grid(row=1+i,column=6)
			entry_child_list.append(e)
			self.entry_list.append(entry_child_list)
			
		return 2+i
	
	def show_content_income_detail(self):
		if self.data == None:
			self.data = []
			return 2
		self.comboxlist_part_No_list = []
		self.entry_list = []
		i = 0
		for i in range(1,len(self.data)+1):
			entry_child_list = []
			Label(self.top,text = str(i),width = 16).grid(row=1+i,column=0)
			part_No_list_tmp = self.part_No_list[:]
			if self.data[i-1][0] not in part_No_list_tmp:
				part_No_list_tmp.append(self.data[i-1][0])
			comboxlist_part_No = self.create_drop_down_list(part_No_list_tmp)
			comboxlist_part_No.bind("<<ComboboxSelected>>",self.refresh)  #绑定事件,下拉列表框被选中时，绑定函数
			comboxlist_part_No.current(part_No_list_tmp.index(self.data[i-1][0]))
			comboxlist_part_No.grid(row=1+i,column=1)
			self.comboxlist_part_No_list.append(comboxlist_part_No)
			e = Entry(self.top)
			e.insert(0,str(self.data[i-1][1]))
			e.grid(row=1+i,column=2)
			entry_child_list.append(e)
			e = Entry(self.top)
			e.insert(0,str(self.data[i-1][2]))
			e.grid(row=1+i,column=3)
			entry_child_list.append(e)
			e = Entry(self.top)
			e.insert(0,str(self.data[i-1][3]))
			e.grid(row=1+i,column=4)
			entry_child_list.append(e)
			e = Entry(self.top)
			e.insert(0,str(self.data[i-1][4]))
			e.grid(row=1+i,column=5)
			entry_child_list.append(e)
			e = Entry(self.top)
			e.insert(0,str(self.data[i-1][5]))
			e.grid(row=1+i,column=6)
			entry_child_list.append(e)
			e = Entry(self.top)
			e.insert(0,str(self.data[i-1][6]))
			e.grid(row=1+i,column=7)
			entry_child_list.append(e)
			e = Entry(self.top)
			e.insert(0,str(self.data[i-1][7]))
			e.grid(row=1+i,column=8)
			entry_child_list.append(e)
			self.entry_list.append(entry_child_list)
			
		return 2+i
	
	def show_content_commit(self,y=2):
		Button(self.top,text = "添加",width = 16,command = self.add_data).grid(row=y,column=2)
		delete_select_list = []
		for i in range(len(self.data)):
			delete_select_list.append(str(i+1))
		delete_select_list.append("删除")
		self.comboxlist_delete_data = self.create_drop_down_list(delete_select_list)
		self.comboxlist_delete_data.bind("<<ComboboxSelected>>",self.delete_data)  #绑定事件,下拉列表框被选中时，绑定函数
		self.comboxlist_delete_data.current(delete_select_list.index("删除"))  
		self.comboxlist_delete_data.grid(row=y,column=3)
		Button(self.top,text = "刷新",width = 16,command = self.refresh).grid(row=y,column=4)
		Button(self.top,text = "保存",width = 16,command = self.save_data).grid(row=y,column=5)
		Button(self.top,text = "返回",width = 16,command = self.back_from_detail).grid(row=y,column=6)
		
	def back_from_detail(self):
		self.top_destory()
		if self.state == "收入":
			self.income_win()
		else:
			self.expenditure_win()
		
	def delete_data(self,*argv):
		try:
			line = int(self.comboxlist_delete_data.get())
		except:
			return
		del self.data[line-1]
		self.top_destory()
		self.show_content_title()
		if self.state == "收入":
			self.show_content_income_item()
			y = self.show_content_income_detail()
		else:
			self.show_content_expenditure_item()
			y = self.show_content_expenditure_detail()
		self.show_content_commit(y)
	
	def add_data(self):
		self.top_destory()
		if self.state == "收入":
			data_child = ["","","","","","","",""]
		else:
			data_child = ["","","","","",""]
		self.data.append(data_child)
		self.show_content_title()
		if self.state == "收入":
			self.show_content_income_item()
			y = self.show_content_income_detail()
		else:
			self.show_content_expenditure_item()
			y = self.show_content_expenditure_detail()
		self.show_content_commit(y)
	
	def collect_income_data(self):
		self.data = []
		i = 0
		for comboxlist_part_No in self.comboxlist_part_No_list:
			data_child = []
			part_No = comboxlist_part_No.get()
			data_child.append(part_No)
			item_type = self.my_database.check_part_No(part_No)
			if item_type == None:
				item_type = self.entry_list[i][0].get()
				if part_No != "" and item_type != "":
					self.my_database.add_part_No(part_No,item_type)
			data_child.append(item_type)
			num = self.entry_list[i][1].get()
			data_child.append(num)
			folding_box_unit_price = self.entry_list[i][2].get()
			data_child.append(folding_box_unit_price)
			try:
				folding_box_total_price = str(Decimal(num)*Decimal(folding_box_unit_price))
			except:
				folding_box_total_price = ""
			data_child.append(folding_box_total_price)
			unit_price = self.entry_list[i][4].get()
			data_child.append(unit_price)
			try:
				total_price = str(Decimal(unit_price)*Decimal(num))
			except:
				total_price = ""
			data_child.append(total_price)
			delivery_note_number = self.entry_list[i][6].get()
			data_child.append(delivery_note_number)
			self.data.append(data_child)
			i += 1
			
	def collect_expenditure_data(self):
		self.data = []
		i = 0
		for comboxlist_name in self.comboxlist_name_list:
			data_child = []
			name = comboxlist_name.get()
			data_child.append(name)
			wages = self.my_database.check_name(name)
			if wages == None:
				wages = self.entry_list[i][0].get()
				if name != "" and wages != "":
					self.my_database.add_name(name,wages)
			data_child.append(wages)
			time = self.entry_list[i][1].get()
			data_child.append(time)
			try:
				amount = str(Decimal(time)*Decimal(wages))
			except:
				amount = ""
			data_child.append(amount)
			subsidy = self.entry_list[i][3].get()
			data_child.append(subsidy)
			information = self.entry_list[i][4].get()
			data_child.append(information)
			self.data.append(data_child)
			i += 1
		
		
	def save_data(self):
		if self.state == "收入":
			self.collect_income_data()
			file_path = os.path.join(self.my_database.path_income,str(self.select_year),str(self.select_month),str(self.select_mday))
			self.my_database.write_income_data(file_path,self.data)
		else:
			self.collect_expenditure_data()
			file_path = os.path.join(self.my_database.path_expenditure,str(self.select_year),str(self.select_month),str(self.select_mday))
			self.my_database.write_expenditure_data(file_path,self.data)
	
	def refresh(self,*argv):
		if self.state == "收入":
			self.collect_income_data()
		else:
			self.collect_expenditure_data()
		self.top_destory()
		self.show_content_title()
		if self.state == "收入":
			self.show_content_income_item()
			y = self.show_content_income_detail()
		else:
			self.show_content_expenditure_item()
			y = self.show_content_expenditure_detail()
		self.show_content_commit(y)
	
	
	def run(self):
		"""开始运行"""
		self.main_win()
		self.top.mainloop()



if __name__ == "__main__":
	my_database = Database()
	my_excel = Excel(my_database)
	#my_database.create_part_No_file()
	my_gui = GUI(my_database,my_excel)
	my_gui.run()
	